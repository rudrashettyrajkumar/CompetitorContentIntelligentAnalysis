"""Full-bundle exports: JSON and a styled multi-sheet xlsx workbook (EPIC-07)."""

from __future__ import annotations

import io
from typing import Annotated

from fastapi import APIRouter, Depends
from fastapi.responses import JSONResponse, StreamingResponse
from sqlalchemy.orm import Session

from app.api.deps import get_session
from app.api.results_service import ResultsService

router = APIRouter(prefix="/api/exports", tags=["exports"])
SessionDep = Annotated[Session, Depends(get_session)]

_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@router.get("/{run_id}.json")
def export_json(run_id: int, session: SessionDep):
    svc = ResultsService(session)
    svc.require_completed_run(run_id)
    return JSONResponse(svc.bundle(run_id))


@router.get("/{run_id}.xlsx")
def export_xlsx(run_id: int, session: SessionDep):
    svc = ResultsService(session)
    svc.require_completed_run(run_id)
    bundle = svc.bundle(run_id)
    workbook = _build_workbook(run_id, bundle)
    buf = io.BytesIO()
    workbook.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type=_XLSX_MIME,
        headers={"Content-Disposition": f'attachment; filename="run-{run_id}.xlsx"'},
    )


def _build_workbook(run_id: int, bundle: dict):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)

    def sheet(name: str, rows: list[dict]) -> None:
        ws = wb.create_sheet(title=name[:31])
        if not rows:
            ws.append(["(no data)"])
            return
        columns: list[str] = []
        for row in rows:
            for key in row:
                if key not in columns:
                    columns.append(key)
        ws.append(columns)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        for row in rows:
            ws.append([_flatten(row.get(col)) for col in columns])
        for i, col in enumerate(columns, start=1):
            width = max(len(col), *(len(str(_flatten(r.get(col)))) for r in rows)) + 2
            ws.column_dimensions[get_column_letter(i)].width = min(width, 60)
        ws.freeze_panes = "A2"

    sheet("Summary", [bundle["summary"]])
    sheet("Posts", bundle["posts"]["items"])
    sheet("Formats", bundle["formats"])
    sheet("Topics", bundle["topics"])
    sheet("CTAs", bundle["ctas"])
    sheet("Keywords", bundle["keywords"])
    sheet("Campaigns", bundle["campaigns"])
    sheet("Profiles", bundle["profiles"])
    sheet("Cross - White Spaces", bundle["cross"].get("white_spaces", []))
    sheet("Cross - Format Opps", bundle["cross"].get("format_opportunities", []))
    sheet("Cross - Keyword Matrix", bundle["cross"].get("keyword_matrix", []))
    sheet("Top Content", _top_content_rows(bundle["top_content"]))
    sheet("Strategy - Pillars", bundle["strategy"].get("pillars", []))
    sheet("Strategy - Formats", bundle["strategy"].get("recommended_formats", []))
    sheet("Opportunities", bundle["opportunities"].get("opportunities", []))
    sheet("Calendar", bundle["calendar"].get("calendar", {}).get("entries", []))
    return wb


def _top_content_rows(top_content: dict) -> list[dict]:
    rows = []
    for item in top_content.get("items", []):
        why = item.get("why", {})
        rows.append(
            {
                "rank": item.get("rank"),
                "competitor": item.get("competitor"),
                "url": item.get("url"),
                "format": item.get("format"),
                "topic": item.get("topic"),
                "engagement_score": item.get("engagement_score"),
                "engagement_rate": item.get("engagement_rate"),
                "why_summary": why.get("summary"),
                "hook": why.get("hook"),
            }
        )
    return rows


def _flatten(value):
    if isinstance(value, (list, tuple)):
        return ", ".join(str(_flatten(v)) for v in value)
    if isinstance(value, dict):
        return "; ".join(f"{k}={_flatten(v)}" for k, v in value.items())
    return value
