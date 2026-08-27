"""EPIC-07 API: results sections, exports, paging, and 404 / 409 error bodies."""

import asyncio
import io
from pathlib import Path

import httpx
import pytest

from app.api.main import create_app

PROJECT_ROOT = Path(__file__).resolve().parents[1]
SAMPLE = PROJECT_ROOT / "data" / "input" / "sample_competitors.xlsx"

_SECTIONS = [
    "formats",
    "topics",
    "ctas",
    "keywords",
    "campaigns",
    "profiles",
    "cross",
    "top-content",
    "strategy",
    "opportunities",
    "calendar",
]


@pytest.fixture
async def client(tmp_path, monkeypatch):
    monkeypatch.setenv("DATABASE_URL", f"sqlite:///{tmp_path}/api.db")
    monkeypatch.setenv("LLM_FAKE_MODE", "true")
    from app.config.settings import get_app_config, get_settings

    get_settings.cache_clear()
    get_app_config.cache_clear()
    app = create_app()
    async with httpx.ASGITransport(app=app) as transport:
        async with app.router.lifespan_context(app):
            async with httpx.AsyncClient(transport=transport, base_url="http://test") as c:
                yield c
    get_settings.cache_clear()
    get_app_config.cache_clear()


async def _wait(client, run_id, timeout=60.0):
    deadline = asyncio.get_event_loop().time() + timeout
    while asyncio.get_event_loop().time() < deadline:
        body = (await client.get(f"/api/runs/{run_id}")).json()
        if body["status"] in ("completed", "failed"):
            return body
        await asyncio.sleep(0.05)
    raise AssertionError("run did not finish")


@pytest.fixture
async def completed_run(client):
    await client.post(
        "/api/competitors/upload",
        files={"file": ("s.xlsx", SAMPLE.read_bytes(), "application/vnd.ms-excel")},
    )
    run_id = (await client.post("/api/runs", json={"period_days": 90})).json()["id"]
    body = await _wait(client, run_id)
    assert body["status"] == "completed", body["error"]
    return run_id


async def test_summary_kpi_block(client, completed_run):
    s = (await client.get(f"/api/results/{completed_run}/summary")).json()
    assert s["competitors_analyzed"] == 5
    assert s["total_posts"] > 0
    assert s["posts_per_week"] > 0
    assert s["top_competitor"] and s["top_topic"] and s["top_format"]
    assert isinstance(s["top_keywords"], list) and len(s["top_keywords"]) <= 10
    assert s["campaign_count"] >= 1


@pytest.mark.parametrize("section", _SECTIONS)
async def test_every_section_returns_200(client, completed_run, section):
    resp = await client.get(f"/api/results/{completed_run}/{section}")
    assert resp.status_code == 200
    assert resp.json() not in (None,)


async def test_sections_have_expected_shape(client, completed_run):
    formats = (await client.get(f"/api/results/{completed_run}/formats")).json()
    assert formats and {"format", "posts", "avg_engagement"} <= set(formats[0])

    cross = (await client.get(f"/api/results/{completed_run}/cross")).json()
    assert "keyword_matrix" in cross and "white_spaces" in cross

    strategy = (await client.get(f"/api/results/{completed_run}/strategy")).json()
    assert 4 <= len(strategy["pillars"]) <= 6

    opps = (await client.get(f"/api/results/{completed_run}/opportunities")).json()
    assert 8 <= len(opps["opportunities"]) <= 12

    cal = (await client.get(f"/api/results/{completed_run}/calendar")).json()
    assert cal["valid"] is True and cal["calendar"]["entries"]


async def test_posts_paging_and_sort(client, completed_run):
    page1 = (await client.get(f"/api/results/{completed_run}/posts?limit=5&offset=0")).json()
    assert page1["limit"] == 5 and len(page1["items"]) == 5
    total = page1["total"]
    assert total > 5

    page2 = (await client.get(f"/api/results/{completed_run}/posts?limit=5&offset=5")).json()
    assert {p["post_id"] for p in page1["items"]}.isdisjoint(p["post_id"] for p in page2["items"])

    desc = (
        await client.get(f"/api/results/{completed_run}/posts?limit=50&sort=engagement_score")
    ).json()
    scores = [p["engagement_score"] for p in desc["items"]]
    assert scores == sorted(scores, reverse=True)

    asc = (
        await client.get(
            f"/api/results/{completed_run}/posts?limit=50&sort=engagement_score&order=asc"
        )
    ).json()
    assert [p["engagement_score"] for p in asc["items"]] == sorted(
        p["engagement_score"] for p in asc["items"]
    )


async def test_posts_filter_by_competitor(client, completed_run):
    all_posts = (await client.get(f"/api/results/{completed_run}/posts?limit=1000")).json()
    cid = all_posts["items"][0]["competitor_id"]
    filtered = (
        await client.get(f"/api/results/{completed_run}/posts?limit=1000&competitor_id={cid}")
    ).json()
    assert filtered["total"] < all_posts["total"]
    assert all(p["competitor_id"] == cid for p in filtered["items"])


async def test_unknown_run_is_404_rfc7807(client):
    resp = await client.get("/api/results/99999/summary")
    assert resp.status_code == 404
    body = resp.json()
    assert body["status"] == 404 and body["title"] and "type" in body


async def test_results_conflict_is_409(client):
    # a run row that never ran -> status 'pending' -> 409 on results
    await client.post(
        "/api/competitors/upload",
        files={"file": ("s.xlsx", SAMPLE.read_bytes(), "application/vnd.ms-excel")},
    )
    from app.config.settings import get_settings
    from app.db.engine import build_engine, build_session_factory
    from app.db.repos import RunRepo

    engine = build_engine(get_settings().database_url)
    session = build_session_factory(engine)()
    run = RunRepo(session).create(period_days=30, adapter="mock")
    session.commit()
    stuck_id = run.id
    session.close()

    resp = await client.get(f"/api/results/{stuck_id}/summary")
    assert resp.status_code == 409
    assert resp.json()["status"] == 409


async def test_export_json_bundle(client, completed_run):
    resp = await client.get(f"/api/exports/{completed_run}.json")
    assert resp.status_code == 200
    bundle = resp.json()
    for key in (
        "summary",
        "posts",
        "formats",
        "campaigns",
        "profiles",
        "cross",
        "strategy",
        "opportunities",
        "calendar",
    ):
        assert key in bundle


async def test_export_xlsx_has_one_sheet_per_section(client, completed_run):
    from openpyxl import load_workbook

    resp = await client.get(f"/api/exports/{completed_run}.xlsx")
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith("application/vnd.openxmlformats")
    wb = load_workbook(io.BytesIO(resp.content))
    for name in ("Summary", "Posts", "Formats", "Campaigns", "Opportunities", "Calendar"):
        assert name in wb.sheetnames
    # row counts line up with the JSON bundle
    bundle = (await client.get(f"/api/exports/{completed_run}.json")).json()
    assert wb["Posts"].max_row - 1 == len(bundle["posts"]["items"])
    assert wb["Opportunities"].max_row - 1 == len(bundle["opportunities"]["opportunities"])


async def test_export_unknown_run_404(client):
    assert (await client.get("/api/exports/424242.json")).status_code == 404
    assert (await client.get("/api/exports/424242.xlsx")).status_code == 404


async def test_competitor_get_and_delete(client):
    up = await client.post(
        "/api/competitors/upload",
        files={"file": ("s.xlsx", SAMPLE.read_bytes(), "application/vnd.ms-excel")},
    )
    cid = up.json()["stored_competitor_ids"][0]
    assert (await client.get(f"/api/competitors/{cid}")).status_code == 200
    assert (await client.delete(f"/api/competitors/{cid}")).status_code == 204
    assert (await client.get(f"/api/competitors/{cid}")).status_code == 404
    assert (await client.delete("/api/competitors/999999")).status_code == 404


async def test_runs_list_and_stage_progress(client, completed_run):
    runs = (await client.get("/api/runs")).json()
    assert any(r["id"] == completed_run for r in runs)
    one = (await client.get(f"/api/runs/{completed_run}")).json()
    assert one["stages"] == ["collect", "classify", "analyze", "map", "strategy", "loop"]
    assert one["status"] == "completed"
    assert (await client.get("/api/runs/999999")).status_code == 404
